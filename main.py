import os
import requests
from urllib.parse import urlparse
from googlesearch import search
import openpyxl
import hashlib
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

class ScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("File Downloader")

        # Set up GUI components
        # Row 1: Select Search Query File
        self.query_file_label = tk.Label(root, text="Select Search Query File:")
        self.query_file_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        self.query_file_button = tk.Button(root, text="Browse", command=self.browse_query_file)
        self.query_file_button.grid(row=0, column=1, padx=10, pady=10)

        self.query_file_entry = tk.Entry(root, width=50)
        self.query_file_entry.grid(row=0, column=2, padx=10, pady=10)

        # Row 2: Select Output Directory
        self.output_dir_label = tk.Label(root, text="Select Main Output Directory:")
        self.output_dir_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")

        self.output_dir_button = tk.Button(root, text="Browse", command=self.browse_output_dir)
        self.output_dir_button.grid(row=1, column=1, padx=10, pady=10)

        self.output_dir_entry = tk.Entry(root, width=50)
        self.output_dir_entry.grid(row=1, column=2, padx=10, pady=10)

        # Row 3: File Extension Selection
        self.extension_label = tk.Label(root, text="Select File Extension:")
        self.extension_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")

        self.extension_var = tk.StringVar(value="pdf")
        self.extension_menu = ttk.Combobox(root, textvariable=self.extension_var, values=["pdf", "docx", "doc", "pptx", "ppt", "csv", "xlsx", "xls", "rtf"])
        self.extension_menu.grid(row=2, column=1, padx=10, pady=10)

        # Row 4: Scrape Button
        self.scrape_button = tk.Button(root, text="Scrape", command=self.start_scraping)
        self.scrape_button.grid(row=3, column=1, padx=10, pady=10)

        # Row 5: Clear Cache Button
        self.clear_cache_button = tk.Button(root, text="Clear Cache", command=self.clear_cache)
        self.clear_cache_button.grid(row=3, column=2, padx=10, pady=10)

        # Status messages (initially hidden)
        self.status_message = tk.Label(root, text="", fg="purple")
        self.success_label = tk.Label(root, text="", fg="green")

    def browse_query_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        self.query_file_entry.delete(0, tk.END)
        self.query_file_entry.insert(0, file_path)

    def browse_output_dir(self):
        directory_path = filedialog.askdirectory()
        self.output_dir_entry.delete(0, tk.END)
        self.output_dir_entry.insert(0, directory_path)

    def start_scraping(self):
        # Validate inputs
        query_file = self.query_file_entry.get()
        output_dir = self.output_dir_entry.get()
        extension = self.extension_var.get()
        
        if not query_file or not output_dir:
            messagebox.showerror("Input Error", "Please select both the search query file and output directory.")
            return
        
        if not extension:
            messagebox.showerror("Input Error", "Please select a file extension.")
            return
        
        # Remove previous success message and display status message
        self.success_label.grid_remove()
        self.status_message.config(text="Scraping started. Please wait until completion.")
        self.status_message.grid(row=4, column=0, columnspan=3, padx=10, pady=10)
        
        # Perform scraping in a separate thread to avoid freezing the GUI
        self.root.after(100, lambda: self.download_files_from_queries(query_file, output_dir, extension))

    def download_files_from_queries(self, excel_file, download_root, file_extension, num_results=10):
        try:
            # Load the Excel file containing search queries
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # Identify the column with the header "queries"
            query_column = None
            for cell in sheet[1]:  # First row contains headers
                if cell.value and cell.value.lower() == 'queries':
                    query_column = cell.column_letter
                    break

            if not query_column:
                raise ValueError("No 'queries' column found in the Excel file.")
            
            # Keep track of downloaded files using a hash cache
            cache_file = 'downloaded_files_cache.txt'
            downloaded_hashes = set()

            # Load existing hashes if cache file exists
            if os.path.exists(cache_file):
                with open(cache_file, 'r') as f:
                    downloaded_hashes = set(f.read().splitlines())

            # Iterate through search queries in the specified column
            for row in sheet.iter_rows(min_row=2):  # Start from the second row to skip headers
                search_query = row[sheet[query_column + '1'].column - 1].value
                if not search_query:
                    continue  # Skip empty rows
                
                # Create a folder named after the search query
                query_folder = os.path.join(download_root, search_query.replace(" ", "_"))
                if not os.path.exists(query_folder):
                    os.makedirs(query_folder)

                # Create or open the Excel file to record download links
                links_file_path = os.path.join(query_folder, 'download_links.xlsx')
                if os.path.exists(links_file_path):
                    wb_links = openpyxl.load_workbook(links_file_path)
                    ws_links = wb_links.active
                else:
                    wb_links = openpyxl.Workbook()
                    ws_links = wb_links.active
                    ws_links.append(["Link", "Downloaded File Name"])

                print(f"\nProcessing query: {search_query}")
                try:
                    search_results = search(search_query + f" filetype:{file_extension}")
                except Exception as e:
                    print(f"Search failed: {e}")
                    self.show_error("Search failed. Check your internet connection or query syntax.")
                    return
                
                count = 0
                for url in search_results:
                    if count >= num_results:
                        break
                    
                    if url.lower().endswith(f".{file_extension}"):
                        try:
                            # Generate a unique hash for the URL to check for duplicates
                            file_hash = hashlib.md5(url.encode()).hexdigest()
                            if file_hash in downloaded_hashes:
                                print(f"Skipping already downloaded: {url}")
                                continue
                            
                            # Get the filename from the URL
                            filename = os.path.basename(urlparse(url).path)
                            file_path = os.path.join(query_folder, filename)
                            
                            # Download the file
                            print(f"Downloading {filename} from {url}...")
                            response = requests.get(url)
                            with open(file_path, 'wb') as file:
                                file.write(response.content)
                            print(f"Saved: {file_path}")
                            
                            # Record the download link and filename
                            ws_links.append([url, filename])
                            wb_links.save(links_file_path)
                            
                            # Cache the file hash and update the cache file
                            downloaded_hashes.add(file_hash)
                            with open(cache_file, 'a') as f:
                                f.write(file_hash + '\n')
                            
                            count += 1

                        except Exception as e:
                            print(f"Failed to download from {url}. Error: {e}")
                            continue

                # Add a delay between queries to avoid triggering CAPTCHA challenges
                time.sleep(5)

            # Scraping finished successfully
            self.show_success(f"Scraping completed. Files saved in: {download_root}")

        except Exception as e:
            self.show_error(str(e))

    def show_success(self, message):
        # Remove status message and display success message
        self.status_message.grid_remove()
        self.success_label.config(text=message)
        self.success_label.grid(row=5, column=0, columnspan=3, padx=10, pady=10)

    def show_error(self, message):
        # Remove status message and display an error message
        self.status_message.grid_remove()
        messagebox.showerror("Error", message)

    def clear_cache(self):
        cache_file = 'downloaded_files_cache.txt'
        if os.path.exists(cache_file):
            os.remove(cache_file)
            messagebox.showinfo("Cache Cleared", "Cache file has been successfully cleared.")
        else:
            messagebox.showinfo("No Cache Found", "No cache file found to clear.")

if __name__ == "__main__":
    root = tk.Tk()
    app = ScraperApp(root)
    root.mainloop()
