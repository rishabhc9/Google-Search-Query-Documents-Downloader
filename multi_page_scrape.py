import os
import requests
from urllib.parse import urlparse
import openpyxl
import hashlib
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from bs4 import BeautifulSoup
import re

class ScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("File Downloader")

        # Set up GUI components
        self.query_file_label = tk.Label(root, text="Select Search Query File:")
        self.query_file_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        self.query_file_button = tk.Button(root, text="Browse", command=self.browse_query_file)
        self.query_file_button.grid(row=0, column=1, padx=10, pady=10)

        self.query_file_entry = tk.Entry(root, width=50)
        self.query_file_entry.grid(row=0, column=2, padx=10, pady=10)

        self.output_dir_label = tk.Label(root, text="Select Main Output Directory:")
        self.output_dir_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")

        self.output_dir_button = tk.Button(root, text="Browse", command=self.browse_output_dir)
        self.output_dir_button.grid(row=1, column=1, padx=10, pady=10)

        self.output_dir_entry = tk.Entry(root, width=50)
        self.output_dir_entry.grid(row=1, column=2, padx=10, pady=10)

        self.extension_label = tk.Label(root, text="Select File Extension:")
        self.extension_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")

        self.extension_var = tk.StringVar(value="pdf")
        self.extension_menu = ttk.Combobox(root, textvariable=self.extension_var, values=["pdf", "docx", "doc", "pptx", "ppt", "csv", "xlsx", "xls", "rtf"])
        self.extension_menu.grid(row=2, column=1, padx=10, pady=10)

        self.pages_label = tk.Label(root, text="Number of Search Pages per Query:")
        self.pages_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")

        self.pages_entry = tk.Entry(root, width=10)
        self.pages_entry.grid(row=3, column=1, padx=10, pady=10)

        self.scrape_button = tk.Button(root, text="Scrape", command=self.start_scraping)
        self.scrape_button.grid(row=4, column=1, padx=10, pady=10)

        self.clear_cache_button = tk.Button(root, text="Clear Cache", command=self.clear_cache)
        self.clear_cache_button.grid(row=5, column=1, padx=10, pady=10)

        self.status_message = tk.Label(root, text="", fg="purple")
        self.success_label = tk.Label(root, text="", fg="green")

    def browse_query_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        self.query_file_entry.delete(0, tk.END)
        self.query_file_entry.insert(0, file_path)

    def browse_output_dir(self):
        directory_path = filedialog.askdirectory()
        self.output_dir_entry.delete(0, tk.END)
        self.output_dir_entry.insert(0, directory_path)

    def start_scraping(self):
        query_file = self.query_file_entry.get()
        output_dir = self.output_dir_entry.get()
        extension = self.extension_var.get()
        num_pages_str = self.pages_entry.get()
        
        if not query_file or not output_dir:
            messagebox.showerror("Input Error", "Please select both the search query file and output directory.")
            return
        
        if not extension:
            messagebox.showerror("Input Error", "Please select a file extension.")
            return
        
        try:
            num_pages = int(num_pages_str)
        except ValueError:
            messagebox.showerror("Input Error", "Please enter a valid number of pages.")
            return
        
        self.status_message.config(text="Scraping started. Please wait for completion.")
        self.status_message.grid(row=6, column=0, columnspan=3, padx=10, pady=10)

        self.root.after(100, lambda: self.download_files_from_queries(query_file, output_dir, extension, num_pages))

    def perform_search(self, search_url):
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }
        try:
            response = requests.get(search_url, headers=headers)
            response.raise_for_status()  # Raise an error for HTTP codes 4xx/5xx
            soup = BeautifulSoup(response.text, 'html.parser')
            urls = [a['href'] for a in soup.find_all('a', href=True) if a['href'].startswith('http')]
            return urls
        except Exception as e:
            print(f"Search failed: {e}")
            return []

    def download_files_from_queries(self, excel_file, download_root, file_extension, num_pages):
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            cache_file = 'downloaded_files_cache.txt'
            downloaded_hashes = set()

            if os.path.exists(cache_file):
                with open(cache_file, 'r') as f:
                    downloaded_hashes = set(f.read().splitlines())

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2
                search_query = row[0]  # Assuming the queries are in the first column
                if not search_query:
                    continue
                
                query_folder = os.path.join(download_root, search_query.replace(" ", "_"))
                if not os.path.exists(query_folder):
                    os.makedirs(query_folder)

                links_file_path = os.path.join(query_folder, 'download_links.xlsx')
                if os.path.exists(links_file_path):
                    wb_links = openpyxl.load_workbook(links_file_path)
                    ws_links = wb_links.active
                else:
                    wb_links = openpyxl.Workbook()
                    ws_links = wb_links.active
                    ws_links.append(["Link", "Downloaded File Name"])

                print(f"\nProcessing query: {search_query}")
                try:
                    for page in range(num_pages):
                        start_index = page * 10
                        search_url = f"https://www.google.com/search?q={search_query}+filetype:{file_extension}&start={start_index}"
                        search_results = self.perform_search(search_url)

                        count = 0
                        for url in search_results:
                            if count >= num_pages * 10:
                                break
                            
                            if url.lower().endswith(f".{file_extension}"):
                                try:
                                    file_hash = hashlib.md5(url.encode()).hexdigest()
                                    if file_hash in downloaded_hashes:
                                        print(f"Skipping already downloaded: {url}")
                                        continue
                                    
                                    filename = os.path.basename(urlparse(url).path)
                                    file_path = os.path.join(query_folder, filename)
                                    
                                    print(f"Downloading {filename} from {url}...")
                                    response = requests.get(url)
                                    with open(file_path, 'wb') as file:
                                        file.write(response.content)
                                    print(f"Saved: {file_path}")
                                    
                                    ws_links.append([url, filename])
                                    wb_links.save(links_file_path)
                                    
                                    downloaded_hashes.add(file_hash)
                                    with open(cache_file, 'a') as f:
                                        f.write(file_hash + '\n')
                                    
                                    count += 1

                                except Exception as e:
                                    print(f"Failed to download from {url}. Error: {e}")
                                    continue

                        time.sleep(5)

                    self.show_success(f"Scraping completed. Files saved in: {download_root}")

                except Exception as e:
                    self.show_error(f"An error occurred during query processing: {str(e)}")

        except Exception as e:
            self.show_error(f"An error occurred: {str(e)}")

    def clear_cache(self):
        try:
            cache_file = 'downloaded_files_cache.txt'
            if os.path.exists(cache_file):
                os.remove(cache_file)
                messagebox.showinfo("Cache Cleared", "Cache file has been cleared successfully.")
            else:
                messagebox.showinfo("Cache Not Found", "No cache file found to clear.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to clear cache: {str(e)}")

    def show_error(self, message):
        self.status_message.grid_remove()
        self.success_label.grid_remove()
        self.status_message.config(text=message, fg="red")
        self.status_message.grid(row=6, column=0, columnspan=3, padx=10, pady=10)

    def show_success(self, message):
        self.status_message.grid_remove()
        self.success_label.config(text=message)
        self.success_label.grid(row=6, column=0, columnspan=3, padx=10, pady=10)

if __name__ == "__main__":
    root = tk.Tk()
    app = ScraperApp(root)
    root.mainloop()
